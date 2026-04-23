"""Excel (.xlsx) extractor backed by ``openpyxl``.

Returns one entry per sheet with cell values flattened to strings. ``openpyxl``
is optional — install via ``pip install datatree-multimodal[xlsx]``.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from .. import EXTRACTOR_VERSIONS
from ..types import ExtractorKind
from . import Extractor, run_blocking

log = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    _HAS_XLSX = True
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None  # type: ignore[assignment]
    _HAS_XLSX = False


class XlsxExtractor(Extractor):
    kind = ExtractorKind.XLSX
    version = EXTRACTOR_VERSIONS["xlsx"]

    async def _extract(
        self,
        file_path: Path,
        options: dict[str, Any],
        warnings: list[str],
    ) -> dict[str, Any]:
        if not _HAS_XLSX:
            raise RuntimeError(
                "openpyxl not installed; install with "
                "`pip install datatree-multimodal[xlsx]`"
            )
        max_rows = int(options.get("max_rows", 0)) or None
        max_cols = int(options.get("max_cols", 0)) or None
        return await run_blocking(
            self._extract_sync, file_path, max_rows, max_cols, warnings
        )

    @staticmethod
    def _extract_sync(
        file_path: Path,
        max_rows: int | None,
        max_cols: int | None,
        warnings: list[str],
    ) -> dict[str, Any]:
        assert load_workbook is not None
        try:
            wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"failed to open xlsx: {exc}") from exc

        sheets_info: list[dict[str, Any]] = []
        try:
            for ws in wb.worksheets:
                rows: list[list[str]] = []
                row_count = 0
                col_count = 0
                try:
                    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        if max_rows is not None and r_idx >= max_rows:
                            warnings.append(
                                f"sheet {ws.title!r}: truncated at {max_rows} rows"
                            )
                            break
                        if max_cols is not None:
                            row = row[:max_cols]
                        cells = ["" if c is None else str(c) for c in row]
                        rows.append(cells)
                        row_count += 1
                        col_count = max(col_count, len(cells))
                except Exception as exc:  # noqa: BLE001
                    warnings.append(f"sheet {ws.title!r}: read failed: {exc}")
                sheets_info.append(
                    {
                        "title": ws.title,
                        "row_count": row_count,
                        "col_count": col_count,
                        "rows": rows,
                    }
                )
        finally:
            wb.close()

        flattened = "\n".join(
            "\t".join(row) for sheet in sheets_info for row in sheet["rows"]
        )
        return {
            "sheet_count": len(sheets_info),
            "char_count": len(flattened),
            "text": flattened,
            "sheets": sheets_info,
        }


__all__ = ["XlsxExtractor"]
